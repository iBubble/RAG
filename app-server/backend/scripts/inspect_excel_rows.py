import openpyxl

wb = openpyxl.load_workbook("/Volumes/SYRAID/RAG_Files/uploads/7e80966d0bf2/新增耕地潜力分析表_任市项目_.xlsx", data_only=True)
ws = wb.active

print("Sheet name:", ws.title)
print("Max row:", ws.max_row)
print("Max col:", ws.max_column)

import openpyxl

wb = openpyxl.load_workbook("/Volumes/SYRAID/RAG_Files/uploads/7e80966d0bf2/新增耕地潜力分析表_任市项目_.xlsx", data_only=True)
ws = wb.active

print("Sheet name:", ws.title)
print("Max row:", ws.max_row)
print("Max col:", ws.max_column)

print("\n--- Values of Row 1 to 30 for All Columns ---")
for r_idx in range(1, 31):
    row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 21)]
    print(f"Row {r_idx}: {row_vals}")

